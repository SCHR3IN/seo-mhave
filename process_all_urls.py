import pandas as pd
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

df = pd.read_csv('/home/egor/Документы/Antigravity_Project/mhave.ru/seo/tasks/m002/url_registry.csv')

def clean_name(val, fallback=""):
    if pd.isna(val) or not str(val).strip() or str(val) == 'nan':
        return fallback
    s = str(val).strip()
    s = re.sub(r'\s+', ' ', s)
    return s

def decline_accusative(name):
    n = name.strip()
    rules = {
        'Косметика': 'косметику',
        'Пищевые добавки': 'пищевые добавки',
        'Аксессуары': 'аксессуары',
        'Средства для лица': 'средства для лица',
        'Средства для тела': 'средства для тела',
        'Уход за лицом': 'средства для ухода за лицом',
        'Уход за телом': 'средства для ухода за телом',
        'Сыворотки для лица': 'сыворотки для лица',
        'Кремы для лица': 'кремы для лица',
        'Очищение кожи': 'средства для очищения кожи',
    }
    if n in rules:
        return rules[n]
    if n.endswith('ка'):
        return n[:-1] + 'ку'
    elif n.endswith('а') and not n.endswith('ста'):
        return n[:-1] + 'у'
    elif n.endswith('ия'):
        return n[:-2] + 'ию'
    return n.lower()

def generate_new_meta(row):
    url = str(row['url'])
    url_type = str(row['url_type'])
    old_title = clean_name(row['title'])
    old_desc = clean_name(row['description'])
    old_h1 = clean_name(row['h1'])
    
    name = old_h1 if old_h1 else old_title
    name = re.sub(r'—.*$', '', name).strip()
    name = re.sub(r' купить.*$', '', name, flags=re.IGNORECASE).strip()
    name = re.sub(r' \|.*$', '', name).strip()
    if not name:
        parts = [p for p in url.split('/') if p]
        name = parts[-1].replace('-', ' ').capitalize() if parts else "MHAVE"
        
    new_h1 = name
    new_title = ""
    new_desc = ""
    action_note = "Оптимизация по ТЗ M008"
    
    acc_name = decline_accusative(name)

    if url in ["https://mhave.ru/", "https://mhave.ru"]:
        new_h1 = "Профессиональная косметика, космецевтика и пищевые добавки Mhave"
        new_title = "Профессиональная косметика и космецевтика MHAVE — интернет-магазин"
        new_desc = "Купить профессиональную косметику, космецевтику и пищевые добавки в интернет-магазине MHAVE. Оригинальная продукция премиум-брендов с доставкой по Москве и всей России."
        action_note = "Главная страница: зафиксированы УТП, H1 и гео-привязка"
        
    elif "/brands/" in url and url_type == "static_page":
        new_h1 = "Бренды"
        new_title = "Бренды премиум косметики и БАД — купить в интернет-магазине MHAVE"
        new_desc = "Официальные бренды профессиональной косметики и БАД в интернет-магазине MHAVE. 100% оригинальная продукция с быстрой доставкой по Москве и России."
        action_note = "Каталог брендов: устранена 2-словная пустышка в Description"
        
    elif url_type == "brand":
        brand_name = name
        new_h1 = brand_name
        new_title = f"Косметика {brand_name} — купить в интернет-магазине MHAVE"
        new_desc = f"Оригинальная косметика бренда {brand_name} в каталоге MHAVE. Официальный сайт продаж. Быстрая доставка по Москве и России, выгодные цены. Покупайте!"
        action_note = f"Страница бренда {brand_name}: грамматически корректная формулировка"
        
    elif url_type == "catalog_section":
        slash_count = url.replace('https://mhave.ru/catalog/', '').strip('/').count('/')
        if slash_count == 0:
            new_title = f"Купить {acc_name} по выгодной цене в интернет-магазине MHAVE"
            new_desc = f"Широкий выбор оригинальной продукции в категории {name}. Купить с быстрой доставкой по Москве и России в интернет-магазине MHAVE. Звоните!"
            action_note = "Корневой раздел: грамматически корректный заголовок в винительном падеже"
        else:
            new_title = f"Купить {acc_name} в Москве по цене от {1500 if 'лиц' in name.lower() else 2400} руб. — MHAVE"
            new_desc = f"Купить {acc_name} в интернет-магазине MHAVE. Оригинальная косметика премиум-класса по цене от {1500 if 'лиц' in name.lower() else 2400} руб. Быстрая доставка по Москве и РФ. Заказывайте!"
            action_note = "Подраздел: грамматическая форма + цена и гео «в Москве»"
            
    elif url_type == "catalog_product":
        prod_name = name
        new_title = f"Купить {prod_name} по выгодной цене в Москве — MHAVE"
        new_desc = f"Предлагаем купить {prod_name} в интернет-магазине MHAVE. 100% оригинал, профессиональный уход. Доставка по Москве и РФ. Оформите заказ!"
        action_note = "Карточка товара: живой коммерческий сниппет"
        
    elif url_type == "promotion":
        new_title = f"{name} — Акции и скидки в интернет-магазине MHAVE"
        new_desc = f"{name}. Официальные скидки и акции на профессиональную косметику и БАДы в магазине MHAVE. Доставка по Москве и России."
        action_note = "Акции: оптимизирован сниппет под рекламные запросы"
        
    elif url_type == "article" or "blog" in url:
        new_title = f"{name} — Читать в блоге MHAVE"
        new_desc = f"{name}. Читайте подробный обзор и советы экспертов в бьюти-блоге интернет-магазина MHAVE."
        action_note = "Статья/Блог: сниппет E-E-A-T"
        
    elif url_type == "catalog_other":
        if "filter" in url:
            new_title = f"Купить {acc_name} в Москве по цене от 1 800 руб. — MHAVE"
            new_desc = f"Широкий выбор оригинальной продукции в категории {name} в интернет-магазине MHAVE. Быстрая доставка по Москве и России. Официальные цены. Заказывайте!"
            action_note = "Smart Filter: разграничение интента"
        else:
            new_title = f"{name} — интернет-магазин MHAVE"
            new_desc = f"{name} в интернет-магазине профессиональной косметики и космецевтики MHAVE. Быстрая доставка по Москве и РФ."
            action_note = "Служебный раздел каталога"
            
    else:
        new_title = f"{name} — интернет-магазин MHAVE"
        new_desc = f"{name} в интернет-магазине профессиональной косметики и космецевтики MHAVE. Подробная информация на официальном сайте."
        action_note = "Служебная страница"
        
    return {
        'url': url,
        'url_type': url_type,
        'old_h1': old_h1,
        'new_h1': new_h1,
        'old_title': old_title,
        'new_title': new_title,
        'old_desc': old_desc,
        'new_desc': new_desc,
        'action_note': action_note
    }

results = [generate_new_meta(row) for _, row in df.iterrows()]

wb_path = '/home/egor/Документы/Antigravity_Project/mhave.ru/seo/tasks/m008/m008_seo_templates_comparison_and_forecast.xlsx'
wb = openpyxl.load_workbook(wb_path)

if "Реестр всех URL (До и После)" in wb.sheetnames:
    del wb["Реестр всех URL (До и После)"]

ws = wb.create_sheet(title="Реестр всех URL (До и После)")

NAVY_HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
WHITE_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
LIGHT_RED_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="EDF2F8", end_color="EDF2F8", fill_type="solid")
LIGHT_GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)
REGULAR_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", size=10, bold=True)
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1F4E78")
SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color="595959")

ws.cell(row=1, column=1, value="Полный реестр всех 235 URL сайта MHAVE.ru: Сравнение мета-тегов ДО и ПОСЛЕ (M008 - Уточненный)").font = TITLE_FONT
ws.cell(row=2, column=1, value="Полный справочник для автоматического и ручного внедрения шаблонов в Битрикс (Задача M009)").font = SUBTITLE_FONT

headers = [
    "№", "URL страницы", "Тип URL", 
    "Старый H1 (ДО)", "Новый H1 (ПОСЛЕ)",
    "Старый Title (ДО)", "Новый Title M008 (ПОСЛЕ)",
    "Старый Description (ДО)", "Новый Description M008 (ПОСЛЕ)",
    "Примечание / Эффект"
]

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col_idx, value=header)
    cell.fill = NAVY_HEADER_FILL
    cell.font = WHITE_HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for idx, r in enumerate(results, 1):
    r_idx = idx + 4
    row_vals = [
        idx, r['url'], r['url_type'],
        r['old_h1'], r['new_h1'],
        r['old_title'], r['new_title'],
        r['old_desc'], r['new_desc'],
        r['action_note']
    ]
    for c_idx, val in enumerate(row_vals, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        cell.font = REGULAR_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if c_idx == 1:
            cell.alignment = Alignment(horizontal="center", vertical="top")
        elif c_idx == 2:
            cell.font = BOLD_FONT
        elif c_idx in [4, 6, 8]:
            cell.fill = LIGHT_RED_FILL
        elif c_idx in [5, 7, 9]:
            cell.fill = LIGHT_BLUE_FILL
        elif c_idx == 10:
            cell.fill = LIGHT_GREEN_FILL

ws.views.sheetView[0].showGridLines = True
col_widths = {'A': 6, 'B': 35, 'C': 18, 'D': 25, 'E': 25, 'F': 35, 'G': 35, 'H': 40, 'I': 40, 'J': 35}
for col_let, w in col_widths.items():
    ws.column_dimensions[col_let].width = w

wb.save(wb_path)
print("Updated Excel file with refined grammatical titles!")
